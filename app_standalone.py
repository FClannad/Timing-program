from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import sqlite3
from datetime import datetime, timedelta
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

if getattr(sys, 'frozen', False):
    os.chdir(sys._MEIPASS)

class CustomFlask(Flask):
    jinja_options = Flask.jinja_options.copy()
    jinja_options.update(dict(
        variable_start_string='[[',
        variable_end_string=']]',
    ))

app = CustomFlask(__name__)
CORS(app)

class Database:
    def __init__(self):
        app_data = os.path.join(os.path.expanduser('~'), 'PinDouTimer')
        os.makedirs(app_data, exist_ok=True)
        self.db_path = os.path.join(app_data, '拼豆店数据.db')
        self.init_db()

    def get_conn(self):
        return sqlite3.connect(self.db_path)

    def init_db(self):
        conn = self.get_conn()
        cursor = conn.cursor()

        # 计时记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS 计时记录 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                座位号 INTEGER,
                套餐名称 TEXT,
                套餐时长 INTEGER,
                套餐价格 REAL,
                开始时间 TEXT,
                结束时间 TEXT,
                使用时长 TEXT,
                顾客信息 TEXT,
                日期 TEXT,
                已收费 INTEGER DEFAULT 0
            )
        ''')

        # 当前计时状态表（新增）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS 当前计时状态 (
                座位号 INTEGER PRIMARY KEY,
                套餐名称 TEXT,
                套餐时长 INTEGER,
                套餐价格 REAL,
                开始时间 TEXT,
                顾客信息 TEXT
            )
        ''')

        # 套餐表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS 套餐配置 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                名称 TEXT UNIQUE,
                时长 INTEGER,
                价格 REAL
            )
        ''')

        # 会员表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS 会员信息 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                姓名 TEXT,
                手机号 TEXT UNIQUE,
                余额 REAL DEFAULT 0,
                积分 INTEGER DEFAULT 0,
                创建时间 TEXT
            )
        ''')

        # 优惠券表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS 优惠券 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                名称 TEXT,
                折扣 REAL,
                满减金额 REAL,
                使用次数 INTEGER DEFAULT 0,
                有效期 TEXT,
                已使用 INTEGER DEFAULT 0
            )
        ''')

        # 初始化默认套餐
        default_packages = [
            ('3小时套餐', 180, 68.0),
            ('日落场套餐', 240, 88.0),
            ('小板套餐', 120, 48.0),
            ('单人套餐', 180, 58.0),
            ('双人套餐', 180, 98.0)
        ]

        for pkg in default_packages:
            cursor.execute('INSERT OR IGNORE INTO 套餐配置 (名称, 时长, 价格) VALUES (?, ?, ?)', pkg)

        conn.commit()
        conn.close()

    def load_active_timers(self):
        """从数据库加载正在计时的座位"""
        conn = self.get_conn()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM 当前计时状态')
        timers = {}
        for row in cursor.fetchall():
            timers[row[0]] = {
                'is_occupied': True,
                'package_name': row[1],
                'package_duration': row[2],
                'package_price': row[3],
                'start_time': row[4],
                'customer_info': row[5]
            }
        conn.close()
        return timers

db = Database()
seats_status = db.load_active_timers()  # 启动时从数据库恢复计时状态

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/seats', methods=['GET'])
def get_seats():
    return jsonify(seats_status)

@app.route('/api/seats/<int:seat_num>/start', methods=['POST'])
def start_timing(seat_num):
    data = request.json
    seats_status[seat_num] = {
        'is_occupied': True,
        'package_name': data['package_name'],
        'package_duration': data['package_duration'],
        'package_price': data['package_price'],
        'customer_info': data.get('customer_info', ''),
        'member_id': data.get('member_id'),
        'coupon_id': data.get('coupon_id'),
        'start_time': datetime.now().isoformat()
    }

    # 持久化到数据库
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO 当前计时状态 (座位号, 套餐名称, 套餐时长, 套餐价格, 开始时间, 顾客信息)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (seat_num, data['package_name'], data['package_duration'], data['package_price'],
          seats_status[seat_num]['start_time'], data.get('customer_info', '')))
    conn.commit()
    conn.close()

    return jsonify({'success': True})

@app.route('/api/seats/<int:seat_num>/stop', methods=['POST'])
def stop_timing(seat_num):
    if seat_num not in seats_status:
        return jsonify({'success': False, 'error': '座位未使用'})

    seat = seats_status[seat_num]
    start_time = datetime.fromisoformat(seat['start_time'])
    end_time = datetime.now()
    duration = end_time - start_time

    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO 计时记录 (座位号, 套餐名称, 套餐时长, 套餐价格, 开始时间, 结束时间, 使用时长,
                             顾客信息, 日期, 已收费)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        seat_num,
        seat['package_name'],
        seat['package_duration'],
        seat['package_price'],
        start_time.strftime('%Y-%m-%d %H:%M:%S'),
        end_time.strftime('%Y-%m-%d %H:%M:%S'),
        str(timedelta(seconds=int(duration.total_seconds()))),
        seat.get('customer_info', ''),
        datetime.now().strftime('%Y-%m-%d'),
        0
    ))
    conn.commit()
    conn.close()

    # 从数据库删除计时状态
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM 当前计时状态 WHERE 座位号 = ?', (seat_num,))
    conn.commit()
    conn.close()

    del seats_status[seat_num]
    return jsonify({
        'success': True,
        'duration': str(timedelta(seconds=int(duration.total_seconds()))),
        'original_price': seat['package_price'],
        'actual_price': seat['package_price']
    })

@app.route('/api/seats/<int:seat_num>/charge', methods=['POST'])
def charge_seat(seat_num):
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('UPDATE 计时记录 SET 已收费 = 1 WHERE id = (SELECT id FROM 计时记录 WHERE 座位号 = ? AND 已收费 = 0 ORDER BY id DESC LIMIT 1)', (seat_num,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/packages', methods=['GET'])
def get_packages():
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM 套餐配置')
    packages = [{'id': row[0], 'name': row[1], 'duration': row[2], 'price': row[3]} for row in cursor.fetchall()]
    conn.close()
    return jsonify(packages)

@app.route('/api/packages', methods=['POST'])
def add_package():
    data = request.json
    conn = db.get_conn()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO 套餐配置 (名称, 时长, 价格) VALUES (?, ?, ?)',
                      (data['name'], data['duration'], data['price']))
        conn.commit()
        package_id = cursor.lastrowid
        conn.close()
        return jsonify({'success': True, 'id': package_id})
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'error': '套餐名称已存在'})

@app.route('/api/packages/<int:pkg_id>', methods=['PUT'])
def update_package(pkg_id):
    data = request.json
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('UPDATE 套餐配置 SET 名称 = ?, 时长 = ?, 价格 = ? WHERE id = ?',
                  (data['name'], data['duration'], data['price'], pkg_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/packages/<int:pkg_id>', methods=['DELETE'])
def delete_package(pkg_id):
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM 套餐配置')
    if cursor.fetchone()[0] <= 1:
        conn.close()
        return jsonify({'success': False, 'error': '至少保留一个套餐'})

    cursor.execute('DELETE FROM 套餐配置 WHERE id = ?', (pkg_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/members', methods=['GET'])
def get_members():
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM 会员信息 ORDER BY id DESC')
    members = [{
        'id': row[0],
        'name': row[1],
        'phone': row[2],
        'balance': row[3],
        'points': row[4],
        'create_time': row[5]
    } for row in cursor.fetchall()]
    conn.close()
    return jsonify(members)

@app.route('/api/members', methods=['POST'])
def add_member():
    data = request.json
    conn = db.get_conn()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO 会员信息 (姓名, 手机号, 余额, 积分, 创建时间) VALUES (?, ?, ?, ?, ?)',
                      (data['name'], data['phone'], data.get('balance', 0), 0, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        member_id = cursor.lastrowid
        conn.close()
        return jsonify({'success': True, 'id': member_id})
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'error': '手机号已存在'})

@app.route('/api/members/<int:member_id>/recharge', methods=['POST'])
def recharge_member(member_id):
    data = request.json
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('UPDATE 会员信息 SET 余额 = 余额 + ? WHERE id = ?', (data['amount'], member_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/coupons', methods=['GET'])
def get_coupons():
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM 优惠券 WHERE 已使用 = 0')
    coupons = [{
        'id': row[0],
        'name': row[1],
        'discount': row[2],
        'reduction': row[3],
        'usage': row[4],
        'expiry': row[5]
    } for row in cursor.fetchall()]
    conn.close()
    return jsonify(coupons)

@app.route('/api/coupons', methods=['POST'])
def add_coupon():
    data = request.json
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO 优惠券 (名称, 折扣, 满减金额, 使用次数, 有效期) VALUES (?, ?, ?, ?, ?)',
                  (data['name'], data.get('discount'), data.get('reduction'), data.get('usage', 1), data['expiry']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/records', methods=['GET'])
def get_records():
    conn = db.get_conn()
    cursor = conn.cursor()

    # 构建查询条件
    conditions = []
    params = []

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    package_name = request.args.get('package_name')

    if start_date and end_date:
        conditions.append('日期 BETWEEN ? AND ?')
        params.extend([start_date, end_date])

    if package_name:
        conditions.append('套餐名称 = ?')
        params.append(package_name)

    where_clause = ' WHERE ' + ' AND '.join(conditions) if conditions else ''

    cursor.execute(f'SELECT * FROM 计时记录{where_clause} ORDER BY id DESC LIMIT 100', params)
    records = [{
        'id': row[0],
        'seat_num': row[1],
        'package_name': row[2],
        'package_duration': row[3],
        'package_price': row[4],
        'start_time': row[5],
        'end_time': row[6],
        'duration': row[7],
        'customer_info': row[8],
        'date': row[9],
        'charged': row[10]
    } for row in cursor.fetchall()]
    conn.close()
    return jsonify(records)

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    conn = db.get_conn()
    cursor = conn.cursor()

    # 获取筛选参数
    conditions = []
    params = []
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    package_name = request.args.get('package_name')

    if start_date and end_date:
        conditions.append('日期 BETWEEN ? AND ?')
        params.extend([start_date, end_date])
    if package_name:
        conditions.append('套餐名称 = ?')
        params.append(package_name)

    where_clause = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    cursor.execute(f'SELECT * FROM 计时记录{where_clause} ORDER BY id DESC', params)
    records = cursor.fetchall()
    conn.close()

    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "计时记录"

    # 表头样式
    header_fill = PatternFill(start_color="FFD4E5", end_color="FFD4E5", fill_type="solid")
    header_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 写入表头
    headers = ['ID', '座位号', '套餐名称', '套餐时长(分钟)', '套餐价格', '开始时间', '结束时间', '实际时长', '顾客信息', '日期', '收费状态']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 写入数据
    for row in records:
        ws.append([
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7],
            row[8], row[9], '已收费' if row[10] else '未收费'
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'计时记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/stats/day', methods=['GET'])
def get_day_stats():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    conn = db.get_conn()
    cursor = conn.cursor()

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 = ? AND 已收费 = 1', (date,))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 = ? AND 已收费 = 0', (date,))
    unpaid = cursor.fetchone()[0] or 0

    cursor.execute('''
        SELECT 套餐名称, COUNT(*), SUM(套餐价格)
        FROM 计时记录
        WHERE 日期 = ?
        GROUP BY 套餐名称
    ''', (date,))
    package_stats = [{'name': row[0], 'count': row[1], 'income': row[2]} for row in cursor.fetchall()]

    cursor.execute('SELECT COUNT(*) FROM 计时记录 WHERE 日期 = ?', (date,))
    total_orders = cursor.fetchone()[0]

    conn.close()
    return jsonify({
        'total_income': total_income,
        'unpaid': unpaid,
        'package_stats': package_stats,
        'total_orders': total_orders
    })

@app.route('/api/stats/month', methods=['GET'])
def get_month_stats():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = db.get_conn()
    cursor = conn.cursor()

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 LIKE ? AND 已收费 = 1', (month + '%',))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 LIKE ? AND 已收费 = 0', (month + '%',))
    unpaid = cursor.fetchone()[0] or 0

    cursor.execute('''
        SELECT 套餐名称, COUNT(*), SUM(套餐价格)
        FROM 计时记录
        WHERE 日期 LIKE ?
        GROUP BY 套餐名称
    ''', (month + '%',))
    package_stats = [{'name': row[0], 'count': row[1], 'income': row[2]} for row in cursor.fetchall()]

    cursor.execute('SELECT COUNT(*) FROM 计时记录 WHERE 日期 LIKE ?', (month + '%',))
    total_orders = cursor.fetchone()[0]

    # 每日收入趋势
    cursor.execute('''
        SELECT 日期, SUM(套餐价格) as income
        FROM 计时记录
        WHERE 日期 LIKE ?
        GROUP BY 日期
        ORDER BY 日期
    ''', (month + '%',))
    daily_data = [{'date': row[0], 'income': row[1]} for row in cursor.fetchall()]

    conn.close()
    return jsonify({
        'total_income': total_income,
        'unpaid': unpaid,
        'package_stats': package_stats,
        'total_orders': total_orders,
        'daily_data': daily_data
    })

@app.route('/api/stats/year', methods=['GET'])
def get_year_stats():
    year = request.args.get('year', str(datetime.now().year))
    conn = db.get_conn()
    cursor = conn.cursor()

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 LIKE ? AND 已收费 = 1', (year + '%',))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT SUM(套餐价格) FROM 计时记录 WHERE 日期 LIKE ? AND 已收费 = 0', (year + '%',))
    unpaid = cursor.fetchone()[0] or 0

    cursor.execute('''
        SELECT 套餐名称, COUNT(*), SUM(套餐价格)
        FROM 计时记录
        WHERE 日期 LIKE ?
        GROUP BY 套餐名称
    ''', (year + '%',))
    package_stats = [{'name': row[0], 'count': row[1], 'income': row[2]} for row in cursor.fetchall()]

    cursor.execute('SELECT COUNT(*) FROM 计时记录 WHERE 日期 LIKE ?', (year + '%',))
    total_orders = cursor.fetchone()[0]

    # 每月收入统计
    cursor.execute('''
        SELECT substr(日期, 6, 2) as month, SUM(套餐价格) as income
        FROM 计时记录
        WHERE 日期 LIKE ?
        GROUP BY substr(日期, 6, 2)
        ORDER BY month
    ''', (year + '%',))
    monthly_data = [{'month': int(row[0]), 'income': row[1]} for row in cursor.fetchall()]

    conn.close()
    return jsonify({
        'total_income': total_income,
        'unpaid': unpaid,
        'package_stats': package_stats,
        'total_orders': total_orders,
        'monthly_data': monthly_data
    })

@app.route('/api/check-timeout', methods=['GET'])
def check_timeout():
    timeout_seats = []
    current_time = datetime.now()

    for seat_num, seat in seats_status.items():
        start_time = datetime.fromisoformat(seat['start_time'])
        elapsed = (current_time - start_time).total_seconds() / 60

        if elapsed >= seat['package_duration']:
            timeout_seats.append({
                'seat_num': seat_num,
                'package_name': seat['package_name'],
                'elapsed': int(elapsed),
                'package_duration': seat['package_duration'],
                'price': seat['package_price']
            })

    return jsonify(timeout_seats)

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    # 简化版：返回CSV格式
    today = datetime.now().strftime('%Y-%m-%d')
    conn = db.get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM 计时记录 WHERE 日期 = ?', (today,))
    records = cursor.fetchall()
    conn.close()

    csv_content = "座位号,套餐,价格,开始时间,结束时间,时长,实付,状态\n"
    for r in records:
        csv_content += f"{r[1]},{r[2]},{r[4]},{r[5]},{r[6]},{r[7]},{r[11]},{'已收费' if r[13] else '未收费'}\n"

    return jsonify({'csv': csv_content})

import webbrowser
import threading
import time

def open_browser():
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5003')

if __name__ == '__main__':
    threading.Thread(target=open_browser, daemon=True).start()
    print('拼豆店计时管理系统正在启动...')
    print('浏览器即将自动打开...')
    app.run(host='127.0.0.1', port=5003, debug=False, use_reloader=False)
